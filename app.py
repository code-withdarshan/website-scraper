"""
Email & City Scraper — Flask backend.
"""
from __future__ import annotations

import csv
import io
import json
import logging
import os
import re
import threading
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse

import openpyxl
import requests
import xlrd
from flask import Flask, Response, jsonify, render_template, request, stream_with_context

from scraper import scrape_url

# ── Logging ────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
log = logging.getLogger("app")

# ── Flask app ──────────────────────────────────────────────────────
app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.jinja_env.auto_reload = True


@app.after_request
def _no_cache(resp):
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"]        = "no-cache"
    resp.headers["Expires"]       = "0"
    return resp


MAX_WORKERS    = 8
MAX_URLS       = 500
DOMAIN_COOLDOWN = 1.0   # seconds between requests to the same domain

# ── Header / URL detection ─────────────────────────────────────────
WEBSITE_COL_RE = re.compile(
    r"^\s*(website|web\s*site|site|url|domain|homepage|web|link)s?\s*$", re.I
)
GSHEET_RE     = re.compile(r"docs\.google\.com/spreadsheets/d/([a-zA-Z0-9_-]+)", re.I)
GSHEET_GID_RE = re.compile(r"[#?&]gid=(\d+)", re.I)
URL_LIKE_RE   = re.compile(
    r"^(https?://)?([a-z0-9][a-z0-9\-]*\.)+[a-z]{2,}(/.*)?$", re.I
)


# ── Per-domain rate limiter ────────────────────────────────────────
class DomainThrottle:
    """One slot per domain — serializes requests so the same site isn't hammered."""
    def __init__(self, cooldown: float):
        self.cooldown = cooldown
        self._last_hit: dict[str, float] = defaultdict(float)
        self._locks:    dict[str, threading.Lock] = defaultdict(threading.Lock)
        self._meta_lock = threading.Lock()

    def acquire(self, url: str) -> None:
        domain = urlparse(url if "://" in url else "https://" + url).netloc.lower()
        with self._meta_lock:
            lock = self._locks[domain]
        with lock:
            elapsed = time.time() - self._last_hit[domain]
            if elapsed < self.cooldown:
                time.sleep(self.cooldown - elapsed)
            self._last_hit[domain] = time.time()


throttle = DomainThrottle(DOMAIN_COOLDOWN)


def _throttled_scrape(url: str) -> dict | None:
    throttle.acquire(url)
    return scrape_url(url)


# ── Smart column detection ─────────────────────────────────────────
def _looks_like_url(v: str) -> bool:
    return bool(URL_LIKE_RE.match(v.strip()))


def _extract_urls_from_rows(rows: list[list[str]]) -> list[str]:
    if not rows:
        return []
    header = rows[0]
    target_col = next(
        (i for i, c in enumerate(header) if WEBSITE_COL_RE.match(c or "")),
        None,
    )
    if target_col is None:
        cols = max(len(r) for r in rows)
        scores = [0] * cols
        for r in rows:
            for i, c in enumerate(r):
                if c and _looks_like_url(c):
                    scores[i] += 1
        if max(scores, default=0) >= 2:
            target_col = scores.index(max(scores))

    urls: list[str] = []
    if target_col is not None:
        start = 1 if WEBSITE_COL_RE.match(header[target_col] or "") else 0
        for r in rows[start:]:
            if target_col < len(r):
                v = (r[target_col] or "").strip()
                if v and _looks_like_url(v):
                    urls.append(v)
    if not urls:
        for r in rows:
            for c in r:
                v = (c or "").strip()
                if v and _looks_like_url(v):
                    urls.append(v)
    return urls


# ── File parsers ───────────────────────────────────────────────────
def _csv_rows(data: bytes) -> list[list[str]]:
    text = data.decode("utf-8-sig", errors="replace")
    return [list(row) for row in csv.reader(io.StringIO(text))]


def _xls_rows(data: bytes) -> list[list[str]]:
    wb = xlrd.open_workbook(file_contents=data)
    out = []
    for s in wb.sheets():
        for r in range(s.nrows):
            out.append([str(s.cell_value(r, c)).strip() for c in range(s.ncols)])
    return out


def _xlsx_rows(data: bytes) -> list[list[str]]:
    wb = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True)
    out = []
    for s in wb.worksheets:
        for r in s.iter_rows(values_only=True):
            out.append([("" if c is None else str(c)).strip() for c in r])
    return out


# ── Google Sheets ──────────────────────────────────────────────────
def _is_gsheet(u: str) -> bool:
    return bool(GSHEET_RE.search(u))


def _fetch_gsheet_csv(url: str) -> bytes | None:
    m = GSHEET_RE.search(url)
    if not m:
        return None
    sid = m.group(1)
    gid = (GSHEET_GID_RE.search(url) or [None, "0"])[1] if GSHEET_GID_RE.search(url) else "0"
    export_url = f"https://docs.google.com/spreadsheets/d/{sid}/export?format=csv&gid={gid}"
    try:
        r = requests.get(export_url, timeout=20, allow_redirects=True)
        r.raise_for_status()
        return r.content
    except Exception:
        return None


# ── URL collection ─────────────────────────────────────────────────
def _collect_urls(request) -> tuple[list[str], str | None]:
    """Returns (urls, error_message)."""
    urls: list[str] = []

    if "file" in request.files:
        f = request.files["file"]
        fn = (f.filename or "").lower()
        data = f.read()
        try:
            if fn.endswith(".csv"):
                urls = _extract_urls_from_rows(_csv_rows(data))
            elif fn.endswith(".xlsx"):
                urls = _extract_urls_from_rows(_xlsx_rows(data))
            elif fn.endswith(".xls"):
                urls = _extract_urls_from_rows(_xls_rows(data))
            else:
                return [], "Unsupported file type. Use CSV or XLS/XLSX."
        except Exception as exc:
            return [], f"Could not parse file: {exc}"
    else:
        body = request.get_json(silent=True) or {}
        raw = body.get("urls", "") or request.form.get("urls", "")
        cands = [u.strip() for u in re.split(r"[,\n\r]+", raw) if u.strip()]
        for c in cands:
            if _is_gsheet(c):
                sheet = _fetch_gsheet_csv(c)
                if sheet is None:
                    return [], (
                        "Could not download the Google Sheet. "
                        "Make sure it's shared as 'Anyone with the link can view'."
                    )
                urls.extend(_extract_urls_from_rows(_csv_rows(sheet)))
            else:
                urls.append(c)

    urls = list(dict.fromkeys(u.strip() for u in urls if u.strip()))[:MAX_URLS]
    return urls, None


# ── Routes ─────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/scrape", methods=["POST"])
def scrape():
    """Synchronous scrape — used when SSE isn't available (older clients)."""
    urls, err = _collect_urls(request)
    if err:
        return jsonify({"error": err}), 400
    if not urls:
        return jsonify({"error": "No website URLs found in your input."}), 400

    results, failed = [], []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs = {ex.submit(_throttled_scrape, u): u for u in urls}
        for f in as_completed(futs):
            u = futs[f]
            try:
                r = f.result()
                if r:
                    results.append(r)
                else:
                    failed.append(u)
            except Exception as exc:
                log.warning("scrape error for %s: %s", u, exc)
                failed.append(u)

    return jsonify(
        {"results": results, "failed": failed,
         "count": len(results), "scanned": len(urls)}
    )


@app.route("/scrape-stream", methods=["POST"])
def scrape_stream():
    """Stream live progress via Server-Sent Events."""
    urls, err = _collect_urls(request)
    if err:
        return jsonify({"error": err}), 400
    if not urls:
        return jsonify({"error": "No website URLs found in your input."}), 400

    def gen():
        yield f"data: {json.dumps({'event':'start','total':len(urls)})}\n\n"
        results, failed = [], []
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
            futs = {ex.submit(_throttled_scrape, u): u for u in urls}
            done = 0
            for f in as_completed(futs):
                u = futs[f]
                done += 1
                try:
                    r = f.result()
                except Exception as exc:
                    log.warning("stream scrape error for %s: %s", u, exc)
                    r = None
                if r:
                    results.append(r)
                    payload = {"event": "row", "result": r, "done": done, "total": len(urls)}
                else:
                    failed.append(u)
                    payload = {"event": "failed", "url": u, "done": done, "total": len(urls)}
                yield f"data: {json.dumps(payload)}\n\n"
        yield f"data: {json.dumps({'event':'end','count':len(results),'failed_count':len(failed),'scanned':len(urls),'failed':failed})}\n\n"

    return Response(
        stream_with_context(gen()),
        mimetype="text/event-stream",
        headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"},
    )


@app.route("/export", methods=["POST"])
def export():
    body = request.get_json(silent=True) or {}
    rows = body.get("results", [])
    fmt  = (body.get("format") or "csv").lower()
    fieldnames = ["url", "emails", "phones", "socials", "city"]

    # Normalize each row to have all fields
    for r in rows:
        for f in fieldnames:
            r.setdefault(f, "")

    if fmt == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append([f.capitalize() for f in fieldnames])
        for r in rows:
            ws.append([r.get(f, "") for f in fieldnames])
        # Auto-width columns
        for col_idx, f in enumerate(fieldnames, start=1):
            longest = max([len(str(r.get(f, ""))) for r in rows] + [len(f)])
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(longest + 2, 60)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="results.xlsx"'},
        )

    # Default: CSV
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=fieldnames)
    w.writeheader()
    w.writerows(rows)
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": 'attachment; filename="results.csv"'},
    )


# ── Entry point ────────────────────────────────────────────────────
def _run_waitress(port: int):
    """Production server (no dev-warning, handles concurrency properly)."""
    from waitress import serve
    log.info("Starting waitress on http://0.0.0.0:%s", port)
    serve(app, host="0.0.0.0", port=port, threads=16)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    try:
        _run_waitress(port)
    except ImportError:
        log.warning("waitress not installed — falling back to Flask dev server")
        app.run(debug=False, host="0.0.0.0", port=port)
