"""
Email & Contact Scraper — Streamlit edition.
Re-uses scraper.py from the Flask version.

Run locally:
    streamlit run streamlit_app.py

Deploy:
    Push this repo to GitHub, then connect at https://share.streamlit.io
    Streamlit Cloud auto-installs requirements.txt and starts the app.
"""
from __future__ import annotations

import csv
import io
import json
import re
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from urllib.parse import urlparse

import openpyxl
import pandas as pd
import requests
import streamlit as st
import xlrd

import db
from scraper import scrape_url, strip_tracking_params


# ── Single SQLite connection shared across reruns ──────────────────
@st.cache_resource(show_spinner=False)
def _get_db():
    """Open one DB connection per app instance. Schema initialized on first call."""
    return db.connect()


# ── Playwright Chromium bootstrap (Streamlit Cloud needs this once) ─
# Chromium isn't bundled with the playwright pip package — it must be
# downloaded with `playwright install chromium` (~300 MB, takes 2-3 minutes
# the first time). Run this in a background thread so the UI loads
# immediately. The JS-rendering fallback gracefully no-ops until Chromium
# is ready; the normal requests-based pipeline doesn't need it at all.
@st.cache_resource(show_spinner=False)
def _bootstrap_chromium_async():
    import threading

    def _install():
        try:
            from playwright.sync_api import sync_playwright
            with sync_playwright() as p:
                p.chromium.launch(headless=True).close()
            return
        except Exception:
            pass
        try:
            subprocess.run(
                [sys.executable, "-m", "playwright", "install", "chromium"],
                check=False, timeout=300,
            )
        except Exception:
            pass

    t = threading.Thread(target=_install, daemon=True)
    t.start()
    return t


_bootstrap_chromium_async()

# ── Page config ────────────────────────────────────────────────────
st.set_page_config(
    page_title="Email & Contact Scraper",
    page_icon="📧",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# Hide default Streamlit chrome
st.markdown(
    """
    <style>
      #MainMenu, footer, header[data-testid="stHeader"] { visibility: hidden; }
      .block-container { padding-top: 1.5rem; padding-bottom: 2rem; }
      .stMetric { background:#f8fafc; border:1px solid #e2e8f0; border-radius:.75rem; padding:.75rem; }
      .hero {
        background: linear-gradient(135deg, #1d4ed8, #1e3a8a);
        color: white; padding: 1.5rem 2rem; border-radius: 1rem;
        margin-bottom: 1.5rem;
      }
      .hero h1 { margin:0; font-size:1.75rem; font-weight:800; }
      .hero p  { margin:.25rem 0 0; opacity:.85; font-size:.9rem; }
      .gsheet-tip {
        background:#ecfdf5; border:1px solid #a7f3d0; color:#065f46;
        border-radius:.5rem; padding:.6rem .85rem; font-size:.82rem;
        margin-bottom:.75rem;
      }
    </style>
    """,
    unsafe_allow_html=True,
)


# ── Session state ──────────────────────────────────────────────────
if "results"     not in st.session_state: st.session_state.results     = []
if "failed"      not in st.session_state: st.session_state.failed      = []
if "history"     not in st.session_state: st.session_state.history     = []
if "source_name" not in st.session_state: st.session_state.source_name = None
# In-flight scrape state — set when a scrape starts, cleared when complete.
# Allows the user to resume after a crash, refresh, or websocket disconnect.
#   {"remaining": [...], "done": [...], "results": [...], "failed": [...],
#    "source_name": str, "total": int}
if "scrape_state" not in st.session_state: st.session_state.scrape_state = None

# ── Constants ──────────────────────────────────────────────────────
# 5 workers strikes a balance between throughput and Streamlit Cloud's
# ~1 GB memory limit. Each worker can hold a Chromium open briefly for the
# JS-rendering fallback; combined with the scraper's render semaphore (max 1
# concurrent Chromium globally) peak memory stays around 600-700 MB.
MAX_WORKERS = 5
MAX_URLS    = 0   # 0 = unlimited; scrape every URL in the input
HIST_MAX    = 5


# Public usage counter (abacus.jasoncameron.dev — free, no auth, no PII)
COUNTER_NS  = "email-scraper-darshan"   # unique to this app
COUNTER_KEY = "scrapes"
COUNTER_URL_BUMP = f"https://abacus.jasoncameron.dev/hit/{COUNTER_NS}/{COUNTER_KEY}"
COUNTER_URL_GET  = f"https://abacus.jasoncameron.dev/get/{COUNTER_NS}/{COUNTER_KEY}"


@st.cache_data(ttl=60, show_spinner=False)
def _get_counter() -> int | None:
    """Fetch current counter value. Cached 60s so we don't hammer the API."""
    try:
        r = requests.get(COUNTER_URL_GET, timeout=3)
        if r.ok:
            return int(r.json().get("value", 0))
    except Exception:
        pass
    return None


def _bump_counter() -> int | None:
    """Increment counter and return new value. Fail silently if offline."""
    try:
        r = requests.get(COUNTER_URL_BUMP, timeout=3)
        if r.ok:
            _get_counter.clear()   # invalidate cache so display refreshes
            return int(r.json().get("value", 0))
    except Exception:
        pass
    return None

WEBSITE_COL_RE = re.compile(
    r"^\s*(website|web\s*site|site|url|domain|homepage|web|link)s?\s*$", re.I
)
GSHEET_RE     = re.compile(r"docs\.google\.com/spreadsheets/d/([a-zA-Z0-9_-]+)", re.I)
GSHEET_GID_RE = re.compile(r"[#?&]gid=(\d+)", re.I)
URL_LIKE_RE   = re.compile(
    r"^(https?://)?([a-z0-9][a-z0-9\-]*\.)+[a-z]{2,}(/.*)?$", re.I
)

# TLDs to skip by default — government, education, non-profit organisations
# Matches: .gov, .gov.uk, .gov.in, .edu, .edu.au, .ac.uk, .org, .org.uk, etc.
SKIP_TLD_RE = re.compile(r"(?:^|\.)(gov|edu|ac|org)(\.[a-z]{2,3})?$", re.IGNORECASE)

# Chinese TLD — .cn or any .X.cn / .cn.X variant
CHINESE_TLD_RE = re.compile(r"\.cn(?:\.[a-z]{2,3})?$|(?:^|\.)cn\.", re.IGNORECASE)

# Major mainland-China cities (English + common Pinyin variants) + HK / Macao
CHINESE_CITIES = {
    "beijing", "peking", "shanghai", "guangzhou", "canton", "shenzhen", "tianjin",
    "chongqing", "wuhan", "chengdu", "nanjing", "hangzhou", "xi'an", "xian",
    "qingdao", "tsingtao", "shenyang", "harbin", "changsha", "dalian", "jinan",
    "kunming", "fuzhou", "xiamen", "amoy", "suzhou", "ningbo", "wuxi", "foshan",
    "dongguan", "zhengzhou", "hefei", "nanchang", "guiyang", "nanning",
    "lanzhou", "yinchuan", "xining", "urumqi", "hohhot", "lhasa", "shantou",
    "zhuhai", "wenzhou", "taiyuan", "shijiazhuang", "changchun", "nantong",
    "yantai", "weifang", "linyi", "tangshan", "baoding", "luoyang", "zibo",
    "haikou", "sanya", "huizhou", "zhongshan", "jiangmen", "putian", "quanzhou",
    "hong kong", "hongkong", "kowloon", "macau", "macao",
}


def _matched_skip_tld(url: str) -> str | None:
    """Return the matched TLD suffix (e.g. '.gov', '.edu.au') if URL should be skipped."""
    try:
        host = urlparse(url if "://" in url else "https://" + url).netloc.lower()
        m = SKIP_TLD_RE.search(host)
        return m.group(0) if m else None
    except Exception:
        return None


def _is_chinese_url(url: str) -> bool:
    """Match .cn and variants like .edu.cn, .com.cn, .gov.cn."""
    try:
        host = urlparse(url if "://" in url else "https://" + url).netloc.lower()
        return bool(CHINESE_TLD_RE.search(host))
    except Exception:
        return False


def _is_chinese_city(city: str) -> bool:
    """Match city names belonging to mainland China, HK, or Macao."""
    if not city:
        return False
    c = re.sub(r"[^\w\s']", "", city).strip().lower()
    return c in CHINESE_CITIES


# ── Smart column / URL extraction (same logic as Flask app.py) ─────
def _looks_like_url(v: str) -> bool:
    return bool(URL_LIKE_RE.match(v.strip()))


def _extract_urls_from_rows(rows):
    if not rows:
        return []
    header = rows[0]
    target_col = next(
        (i for i, c in enumerate(header) if WEBSITE_COL_RE.match(c or "")),
        None,
    )
    if target_col is None:
        cols = max(len(r) for r in rows)
        scores = [0] * cols
        for r in rows:
            for i, c in enumerate(r):
                if c and _looks_like_url(c):
                    scores[i] += 1
        if max(scores, default=0) >= 2:
            target_col = scores.index(max(scores))

    urls = []
    if target_col is not None:
        start = 1 if WEBSITE_COL_RE.match(header[target_col] or "") else 0
        for r in rows[start:]:
            if target_col < len(r):
                v = (r[target_col] or "").strip()
                if v and _looks_like_url(v):
                    urls.append(v)
    if not urls:
        for r in rows:
            for c in r:
                v = (c or "").strip()
                if v and _looks_like_url(v):
                    urls.append(v)
    return urls


def _csv_rows(data: bytes):
    text = data.decode("utf-8-sig", errors="replace")
    return [list(row) for row in csv.reader(io.StringIO(text))]


def _xlsx_rows(data: bytes):
    wb = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True)
    out = []
    for s in wb.worksheets:
        for r in s.iter_rows(values_only=True):
            out.append([("" if c is None else str(c)).strip() for c in r])
    return out


def _xls_rows(data: bytes):
    wb = xlrd.open_workbook(file_contents=data)
    out = []
    for s in wb.sheets():
        for r in range(s.nrows):
            out.append([str(s.cell_value(r, c)).strip() for c in range(s.ncols)])
    return out


def _fetch_gsheet_csv(url: str) -> tuple[bytes | None, str | None]:
    """Returns (csv_bytes, sheet_title). Sheet title is read from Content-Disposition."""
    m = GSHEET_RE.search(url)
    if not m:
        return None, None
    sid = m.group(1)
    gid = (GSHEET_GID_RE.search(url) or [None, "0"])[1] if GSHEET_GID_RE.search(url) else "0"
    export_url = f"https://docs.google.com/spreadsheets/d/{sid}/export?format=csv&gid={gid}"
    try:
        r = requests.get(export_url, timeout=20, allow_redirects=True)
        r.raise_for_status()
        # Parse filename from Content-Disposition: 'attachment; filename="Sheet Name - Tab.csv"'
        title = None
        cd = r.headers.get("Content-Disposition", "")
        m2 = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^";]+)"?', cd, re.I)
        if m2:
            raw = m2.group(1).strip()
            try:
                from urllib.parse import unquote as _uq
                raw = _uq(raw)
            except Exception:
                pass
            title = re.sub(r"\.csv$", "", raw, flags=re.I).strip()
            title = _dedupe_gsheet_title(title)
        return r.content, title
    except Exception:
        return None, None


def _dedupe_gsheet_title(title: str) -> str:
    """
    Google exports as "Spreadsheet Name - Tab Name.csv". When both names are
    identical we get "Name - Name", which becomes a doubled download filename.
    Collapse such duplicates.
    """
    # Try common separators Google uses (" - ", " — ", " – ")
    for sep in (" - ", " — ", " – "):
        if sep in title:
            parts = [p.strip() for p in title.split(sep) if p.strip()]
            if len(parts) >= 2:
                # All parts identical → keep one
                if all(p.lower() == parts[0].lower() for p in parts):
                    return parts[0]
                # Adjacent duplicates ("X - X - Y") → dedupe consecutively
                deduped = [parts[0]]
                for p in parts[1:]:
                    if p.lower() != deduped[-1].lower():
                        deduped.append(p)
                return sep.join(deduped)
    return title


def _collect_urls(text_input: str, uploaded_file) -> tuple[list[str], str | None, str | None]:
    """
    Returns (urls, error, source_name).
    `source_name` is set when the input was a Google Sheet or uploaded file,
    so exports can be named after the source.
    """
    urls: list[str] = []
    source_name: str | None = None

    if uploaded_file is not None:
        fn = uploaded_file.name
        data = uploaded_file.read()
        # Strip extension for the source name
        source_name = re.sub(r"\.(csv|xls|xlsx)$", "", fn, flags=re.I).strip()
        fn_lower = fn.lower()
        try:
            if fn_lower.endswith(".csv"):
                urls = _extract_urls_from_rows(_csv_rows(data))
            elif fn_lower.endswith(".xlsx"):
                urls = _extract_urls_from_rows(_xlsx_rows(data))
            elif fn_lower.endswith(".xls"):
                urls = _extract_urls_from_rows(_xls_rows(data))
            else:
                return [], "Unsupported file type.", None
        except Exception as exc:
            return [], f"Could not parse file: {exc}", None
    if text_input:
        # Accept commas, semicolons, whitespace (including newlines & tabs)
        # as separators between URLs / Google Sheet links
        cands = [u.strip() for u in re.split(r"[,\n\r\t;\s]+", text_input) if u.strip()]

        sheet_titles: list[str] = []
        failed_sheets: list[str] = []
        gsheet_count = 0

        for c in cands:
            if GSHEET_RE.search(c):
                gsheet_count += 1
                sheet_bytes, title = _fetch_gsheet_csv(c)
                if sheet_bytes is None:
                    # Skip this one but keep going with the rest
                    failed_sheets.append(c)
                    continue
                urls.extend(_extract_urls_from_rows(_csv_rows(sheet_bytes)))
                if title:
                    sheet_titles.append(title)
            else:
                urls.append(c)

        # Build source name
        if sheet_titles:
            if len(sheet_titles) == 1:
                source_name = sheet_titles[0]
            else:
                # Use first sheet's title + " (+N more)"
                source_name = f"{sheet_titles[0]} (+{len(sheet_titles) - 1} more)"

        # Hard fail only if EVERY Google Sheet failed and there are no other URLs
        if gsheet_count > 0 and not urls and failed_sheets:
            return [], (
                f"All {gsheet_count} Google Sheet link"
                f"{'s' if gsheet_count != 1 else ''} failed to download. "
                "Make sure each sheet is shared as 'Anyone with the link can view'."
            ), None

        # Soft warning when some succeeded, some failed — pass it back via a magic
        # marker the caller can detect (we still want to scrape what we got)
        if failed_sheets:
            # Stash the failure count in source_name suffix so caller can show a hint;
            # but only do it if source_name is otherwise set. Otherwise just log silently.
            pass   # we surface this through st.warning below

        # Save the failure list on a module-level holder so the caller can read it
        global _LAST_FAILED_SHEETS
        _LAST_FAILED_SHEETS = failed_sheets

    urls = [u.strip() for u in urls if u.strip()]
    return urls, None, source_name


# Tracker for partial Google Sheet failures (read by the caller after _collect_urls)
_LAST_FAILED_SHEETS: list[str] = []


def _safe_filename(name: str) -> str:
    """Strip filesystem-unsafe chars and collapse whitespace for a filename."""
    name = re.sub(r"[^\w\s.\-]", "", name).strip()
    name = re.sub(r"\s+", "-", name)
    # Collapse runs of hyphens that come from " - " separators
    name = re.sub(r"-{2,}", "-", name)
    return (name or "scrape-results")[:80]


def _normalize_url(u: str) -> str:
    """Lowercase host, drop trailing slash, www, and tracking params for dedup."""
    u = strip_tracking_params(u)
    parsed = urlparse(u if "://" in u else "https://" + u)
    host = (parsed.netloc or u).lower()
    if host.startswith("www."):
        host = host[4:]
    return (host + (parsed.path or "")).rstrip("/")


def _has_gmail(result: dict) -> bool:
    """True if any email in this row ends with @gmail.com or @googlemail.com."""
    emails = (result.get("emails") or "").lower()
    return "@gmail.com" in emails or "@googlemail.com" in emails


def _sort_gmail_first(results: list[dict]) -> list[dict]:
    """Stable sort: rows whose emails include any Gmail address go to the top."""
    return sorted(results, key=lambda r: 0 if _has_gmail(r) else 1)


def _estimate_time(unique_urls: list[str]) -> tuple[int, int]:
    """
    Rough wall-clock estimate in seconds + unique domain count.
    Assumes ~4s per URL, MAX_WORKERS in parallel, throttled to 1 req/sec per domain.
    """
    if not unique_urls:
        return 0, 0
    domains_count: dict[str, int] = {}
    for u in unique_urls:
        d = _normalize_url(u)
        domains_count[d] = domains_count.get(d, 0) + 1
    max_per_domain = max(domains_count.values())
    n = len(unique_urls)
    # Two bounding factors:
    #   a) worker capacity: n / MAX_WORKERS * 4s per URL
    #   b) per-domain serial scheduling: max_per_domain * 4s
    est = max(n / MAX_WORKERS * 4, max_per_domain * 4, 3)
    return int(est), len(domains_count)


# ── Scrape state — backed by SQLite for crash / disconnect survival ──
def _refresh_scrape_state_from_db() -> None:
    """Hydrate session_state.scrape_state from the most recent unfinished scrape.

    Called at the top of each rerun so resume works across page refreshes, OOM
    kills, and Cloud container restarts (anything that wipes session_state).
    """
    if st.session_state.scrape_state:
        return  # Already populated this rerun
    conn = _get_db()
    active = db.get_active_scrape(conn)
    if not active or active["pending"] == 0:
        return
    sid = active["id"]
    st.session_state.scrape_state = {
        "id":           sid,
        "total":        active["total"],
        "done_count":   active["done"] + active["failed"],
        "remaining":    db.get_remaining_urls(conn, sid),
        "results":      db.get_results(conn, sid),
        "failed":       db.get_failed_urls(conn, sid),
        "source_name":  active["source_name"],
        "skip_chinese": active["skip_chinese"],
    }


def _clear_scrape_state() -> None:
    """Discard the active scrape (deletes its DB rows + clears session_state)."""
    state = st.session_state.scrape_state
    if state and state.get("id"):
        try:
            db.delete_scrape(_get_db(), state["id"])
        except Exception:
            pass
    st.session_state.scrape_state = None


_refresh_scrape_state_from_db()


def _run_scrape(urls_to_process: list[str]) -> None:
    """Process URLs and write each result to the DB as it completes.

    The DB is the source of truth — if the process dies mid-scrape, refreshing
    the page brings the resume banner back with everything already done.
    """
    state = st.session_state.scrape_state
    sid = state["id"]
    total = state["total"]
    done_count = state["done_count"]
    conn = _get_db()

    progress = st.progress(
        done_count / total if total else 0,
        text=f"Scraping {done_count} / {total}…",
    )
    ticker = st.empty()
    ticker_lines: list[str] = []

    try:
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
            futures = {ex.submit(scrape_url, u): u for u in urls_to_process}
            for f in as_completed(futures):
                u = futures[f]
                try:
                    r = f.result()
                except Exception:
                    r = None
                if r:
                    state["results"].append(r)
                    db.mark_url_done(conn, sid, u, r)
                    ticker_lines.insert(0, f"✅ **{u}** — {r['emails'][:80]}")
                else:
                    state["failed"].append(u)
                    db.mark_url_failed(conn, sid, u)
                    ticker_lines.insert(0, f"❌ **{u}** — no email")
                try:
                    state["remaining"].remove(u)
                except ValueError:
                    pass
                done_count += 1
                state["done_count"] = done_count

                progress.progress(
                    done_count / total,
                    text=f"Scraping {done_count} / {total}…",
                )
                ticker.markdown("\n".join(ticker_lines[:15]))
    finally:
        progress.empty()
        ticker.empty()


# ── Sidebar filter (defined early so it's available before scrape logic) ──
st.sidebar.markdown("### ⚙️ Filters")
skip_noncommercial = st.sidebar.checkbox(
    "Skip non-commercial domains",
    value=True,
    help="When ON, .gov / .edu / .ac / .org (and country variants like .gov.uk, "
         ".edu.au) are filtered out before scraping. Turn OFF to scrape them too.",
)
skip_chinese = st.sidebar.checkbox(
    "Skip Chinese websites",
    value=True,
    help="Drops URLs ending in .cn (and .edu.cn, .gov.cn, .com.cn etc.) BEFORE scraping, "
         "and removes any result whose detected City is a known mainland-China / HK / Macao city.",
)
st.sidebar.markdown("---")


# ── Hero ───────────────────────────────────────────────────────────
st.markdown(
    """
    <div class="hero">
      <h1>📧 Email &amp; Contact Scraper</h1>
      <p>Pulls emails, phone numbers, social links &amp; city info from any website's footer, contact, or about page.</p>
    </div>
    """,
    unsafe_allow_html=True,
)

# ── Resume banner (only shown when an unfinished scrape exists) ────
_pending = st.session_state.scrape_state
resume_clicked = False
if _pending and _pending.get("remaining"):
    _done_n = _pending["done_count"]
    _total_n = _pending["total"]
    _remaining_n = len(_pending["remaining"])
    st.warning(
        f"⏸️ **Unfinished scrape detected.** "
        f"{_done_n} of {_total_n} URLs completed — **{_remaining_n} remaining**. "
        "You can resume where you left off, download what you have so far, "
        "or discard and start over."
    )
    rc1, rc2, rc3, _ = st.columns([1.2, 1.4, 1.1, 4.3])
    with rc1:
        resume_clicked = st.button(
            "▶️ Resume scrape",
            type="primary",
            use_container_width=True,
            key="resume_btn",
        )
    with rc2:
        remaining_csv = "Website\n" + "\n".join(_pending["remaining"])
        st.download_button(
            "⬇ Remaining URLs (.csv)",
            data=remaining_csv,
            file_name="remaining-urls.csv",
            mime="text/csv",
            use_container_width=True,
            key="dl_remaining_btn",
        )
    with rc3:
        if st.button("🗑 Discard", use_container_width=True, key="discard_btn"):
            _clear_scrape_state()
            st.rerun()
    st.markdown("---")


# ── Input section ──────────────────────────────────────────────────
tab_paste, tab_upload, tab_filter = st.tabs(
    ["📝 Paste URLs", "📂 Upload File", "🧹 Filter Only"]
)

text_input = ""
uploaded_file = None

with tab_paste:
    st.markdown(
        '<div class="gsheet-tip"><b>Tip:</b> Paste one or <b>multiple Google Sheet URLs</b> '
        '(comma- or newline-separated) — the app finds the <code>Website</code> column '
        'in each and combines all URLs into a single scrape. '
        'Each sheet must be set to <em>"Anyone with the link can view"</em>.</div>',
        unsafe_allow_html=True,
    )
    text_input = st.text_area(
        label="Enter URLs (one per line or comma-separated) or one or more Google Sheet links",
        height=200,
        placeholder=(
            "https://example.com\n"
            "businesssite.com, anothersite.co.uk\n\n"
            "Or one or more Google Sheet links — comma- or newline-separated:\n"
            "https://docs.google.com/spreadsheets/d/AAA.../edit#gid=0,\n"
            "https://docs.google.com/spreadsheets/d/BBB.../edit#gid=123"
        ),
        label_visibility="visible",
        key="text_input",
    )

with tab_upload:
    uploaded_file = st.file_uploader(
        "Upload a CSV / XLS / XLSX file",
        type=["csv", "xls", "xlsx"],
        accept_multiple_files=False,
    )
    if uploaded_file:
        st.success(f"📄 {uploaded_file.name} ready to scrape")

with tab_filter:
    st.markdown(
        '<div class="gsheet-tip"><b>Filter only — no scraping.</b> Paste rows '
        '(URLs, emails or any text) <i>or</i> Google Sheet links. The tool removes '
        '<b>duplicates</b> and any line containing a <code>.edu</code> or '
        '<code>.org</code> domain, then gives you the cleaned list to download.</div>',
        unsafe_allow_html=True,
    )
    filter_input = st.text_area(
        label="Paste data to filter (one entry per line) or Google Sheet link(s)",
        height=220,
        placeholder=(
            "example.com\n"
            "harvard.edu       ← will be removed\n"
            "wikipedia.org     ← will be removed\n"
            "example.com       ← duplicate, will be removed\n\n"
            "Or paste a Google Sheet URL to pull its Website column."
        ),
        key="filter_input",
    )
    case_insensitive = st.checkbox(
        "Treat values case-insensitively when comparing",
        value=True,
        key="filter_case_insensitive",
    )
    load_clicked = st.button(
        "📥 Load data", type="secondary", key="filter_load_btn"
    )

    # ── Stage 1: load pasted input into a DataFrame and cache it ──
    if load_clicked:
        failed_sheets: list[str] = []
        sheet_frames: list[pd.DataFrame] = []
        plain_lines: list[str] = []

        cands = [
            s.strip() for s in re.split(r"[\n\r,;\t]+", filter_input or "") if s.strip()
        ]
        if not cands:
            st.warning("Paste some data first.")
            st.session_state.pop("filter_loaded_df", None)
        else:
            for c in cands:
                if GSHEET_RE.search(c):
                    sheet_bytes, _title = _fetch_gsheet_csv(c)
                    if sheet_bytes is None:
                        failed_sheets.append(c)
                        continue
                    try:
                        df = pd.read_csv(
                            io.BytesIO(sheet_bytes),
                            dtype=str,
                            keep_default_na=False,
                        )
                        sheet_frames.append(df)
                    except Exception as exc:
                        st.error(f"Could not parse sheet: {exc}")
                else:
                    plain_lines.append(c)

            if failed_sheets:
                st.warning(
                    f"⚠️ {len(failed_sheets)} Google Sheet link(s) could not be "
                    "downloaded — make sure they are shared as 'Anyone with the link can view'."
                )

            if sheet_frames:
                full_df = pd.concat(sheet_frames, ignore_index=True, sort=False).fillna("")
                if plain_lines:
                    extra = pd.DataFrame({full_df.columns[0]: plain_lines})
                    full_df = pd.concat([full_df, extra], ignore_index=True, sort=False).fillna("")
            elif plain_lines:
                full_df = pd.DataFrame({"Value": plain_lines})
            else:
                full_df = pd.DataFrame()

            st.session_state["filter_loaded_df"] = full_df

    # ── Stage 2: show column picker + filter on cached df ──
    loaded_df: pd.DataFrame = st.session_state.get("filter_loaded_df", pd.DataFrame())
    if not loaded_df.empty:
        st.markdown(f"**Loaded:** {len(loaded_df)} rows · {len(loaded_df.columns)} columns")

        # Auto-detect likely dedup columns (Email > Phone > Website)
        _PREFERRED_RE = re.compile(r"(?i)\b(email|e[-_ ]?mail|phone|mobile|website|url|domain)\b")
        default_dedup_cols = [c for c in loaded_df.columns if _PREFERRED_RE.search(str(c))]
        if not default_dedup_cols:
            default_dedup_cols = list(loaded_df.columns)

        dedup_cols = st.multiselect(
            "Dedupe based on these column(s) — rows that share the same value(s) "
            "in ALL selected columns will be treated as duplicates and only one is kept.",
            options=list(loaded_df.columns),
            default=default_dedup_cols,
            key="filter_dedup_cols",
            help="E.g. pick only 'Email' to keep one row per email address, regardless of "
                 "other column differences like Last Contacted date.",
        )

        remove_edu_org = st.checkbox(
            "Also remove rows containing .edu / .org domains",
            value=True,
            key="filter_remove_edu_org",
        )

        run_filter = st.button(
            "🧹 Apply filter", type="primary", key="filter_apply_btn"
        )

        if run_filter:
            if not dedup_cols:
                st.warning("Pick at least one column to dedupe on.")
            else:
                EDU_ORG_RE = re.compile(
                    r"(?:^|\.|@)(edu|org)(\.[a-z]{2,3})?(?:$|/|\?|#|\s|,|;|\Z)", re.I
                )

                seen_keys: set[tuple] = set()
                dup_count = 0
                blank_key_count = 0
                edu_org_count = 0
                kept_rows: list[dict] = []

                for _idx, row in loaded_df.iterrows():
                    if remove_edu_org:
                        joined = " ".join(str(v) for v in row.values)
                        if EDU_ORG_RE.search(joined):
                            edu_org_count += 1
                            continue

                    # Build dedup key from the chosen columns
                    key_parts = []
                    for col in dedup_cols:
                        v = str(row.get(col, "")).strip()
                        if case_insensitive:
                            v = v.lower()
                        key_parts.append(v)
                    key = tuple(key_parts)

                    # If every selected column is blank, treat as "no key" and keep
                    if all(p == "" for p in key_parts):
                        blank_key_count += 1
                        kept_rows.append(row.to_dict())
                        continue

                    if key in seen_keys:
                        dup_count += 1
                        continue
                    seen_keys.add(key)
                    kept_rows.append(row.to_dict())

                kept_df = (
                    pd.DataFrame(kept_rows, columns=loaded_df.columns)
                    if kept_rows
                    else pd.DataFrame(columns=loaded_df.columns)
                )

                total_in = len(loaded_df)
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Input rows", total_in)
                c2.metric("Kept", len(kept_df))
                c3.metric("Duplicates removed", dup_count)
                c4.metric(".edu / .org removed", edu_org_count)
                if blank_key_count:
                    st.caption(
                        f"ℹ️ {blank_key_count} row(s) had blank values in all dedup "
                        "columns and were kept as-is."
                    )

                if not kept_df.empty:
                    st.markdown("**✅ Cleaned data**")
                    st.dataframe(kept_df, use_container_width=True, height=320)
                    ts = datetime.now().strftime("%Y%m%d-%H%M%S")

                    csv_bytes = kept_df.to_csv(index=False).encode("utf-8")

                    xlsx_buf = io.BytesIO()
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Filtered"
                    ws.append(list(kept_df.columns))
                    for _, r in kept_df.iterrows():
                        ws.append([str(v) for v in r.values])
                    wb.save(xlsx_buf)

                    dl1, dl2 = st.columns(2)
                    with dl1:
                        st.download_button(
                            "⬇️ Download as .csv",
                            data=csv_bytes,
                            file_name=f"filtered-{ts}.csv",
                            mime="text/csv",
                            use_container_width=True,
                        )
                    with dl2:
                        st.download_button(
                            "⬇️ Download as .xlsx",
                            data=xlsx_buf.getvalue(),
                            file_name=f"filtered-{ts}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )
                else:
                    st.info("Nothing left after filtering.")

col_a, col_b = st.columns([1, 5])
with col_a:
    scrape_clicked = st.button("🔍 Scrape", type="primary", use_container_width=True)


# ── Scrape ─────────────────────────────────────────────────────────
def _finalize_scrape() -> None:
    """Called once the scrape_state has no remaining URLs — applies post-processing,
    moves results to session_state.results, saves history, and clears scrape_state."""
    state = st.session_state.scrape_state
    if not state:
        return
    results = list(state["results"])
    failed  = list(state["failed"])
    source_name = state.get("source_name")
    skip_chinese_flag = state.get("skip_chinese", False)

    if skip_chinese_flag:
        before_n = len(results)
        results = [r for r in results if not _is_chinese_city(r.get("city", ""))]
        dropped = before_n - len(results)
        if dropped:
            st.info(
                f"🇨🇳 Dropped **{dropped}** result"
                f"{'s' if dropped != 1 else ''} after scraping — city detected "
                "as mainland China / Hong Kong / Macao."
            )

    results = _sort_gmail_first(results)
    gmail_count = sum(1 for r in results if _has_gmail(r))
    if gmail_count:
        st.success(
            f"✉️ {gmail_count} site{'s' if gmail_count != 1 else ''} "
            f"with Gmail address{'es' if gmail_count != 1 else ''} sorted to the top."
        )

    st.session_state.results = results
    st.session_state.failed  = failed
    st.session_state.source_name = source_name
    _bump_counter()
    st.session_state.history.insert(0, {
        "when":        datetime.now().strftime("%Y-%m-%d %H:%M"),
        "count":       len(results),
        "failed":      len(failed),
        "results":     results,
        "failed_urls": failed,
        "source_name": source_name,
    })
    st.session_state.history = st.session_state.history[:HIST_MAX]
    # Mark the DB row as finished (but keep it for history browsing)
    if state.get("id"):
        try:
            db.finish_scrape(_get_db(), state["id"])
        except Exception:
            pass
    st.session_state.scrape_state = None


# ── Path 1: Resume an in-flight scrape ──────────────────────────────
if resume_clicked and st.session_state.scrape_state:
    remaining = list(st.session_state.scrape_state["remaining"])
    if remaining:
        st.session_state.source_name = st.session_state.scrape_state.get("source_name")
        _run_scrape(remaining)
        if not st.session_state.scrape_state["remaining"]:
            _finalize_scrape()
            st.rerun()


# ── Path 2: Fresh scrape from input ─────────────────────────────────
if scrape_clicked:
    # If there's an unfinished scrape, warn the user — don't silently overwrite
    if st.session_state.scrape_state and st.session_state.scrape_state.get("remaining"):
        st.error(
            "There is an unfinished scrape above. Click **Resume scrape** to "
            "continue, or **Discard** to start a new one."
        )
    else:
        raw_urls, err, source_name = _collect_urls(text_input, uploaded_file)
        # Surface any Google Sheet failures from the most recent _collect_urls call
        if _LAST_FAILED_SHEETS:
            st.warning(
                f"⚠️ {len(_LAST_FAILED_SHEETS)} Google Sheet link"
                f"{'s' if len(_LAST_FAILED_SHEETS) != 1 else ''} could not be downloaded "
                "(skipped). Make sure each is shared as 'Anyone with the link can view'."
            )
        if err:
            st.error(err)
        elif not raw_urls:
            st.error("No URLs detected in your input.")
        else:
            st.session_state.source_name = source_name
            # ── Dedup + cap ──────────────────────────────────────────
            # Normalize for dedup (treat www / non-www / trailing slash as same)
            seen: dict[str, str] = {}
            for u in raw_urls:
                key = _normalize_url(u)
                seen.setdefault(key, u)
            urls = list(seen.values())

            dupes = len(raw_urls) - len(urls)

            # ── Filter out .gov / .edu / .org / .ac if toggle is on ─
            skipped_tld: list[tuple[str, str]] = []
            if skip_noncommercial:
                kept = []
                for u in urls:
                    tld = _matched_skip_tld(u)
                    if tld:
                        skipped_tld.append((u, tld))
                    else:
                        kept.append(u)
                urls = kept

            # ── Filter out Chinese URLs (.cn variants) if toggle is on ─
            skipped_chinese_urls: list[str] = []
            if skip_chinese:
                kept = []
                for u in urls:
                    if _is_chinese_url(u):
                        skipped_chinese_urls.append(u)
                    else:
                        kept.append(u)
                urls = kept

            capped = 0
            if MAX_URLS and len(urls) > MAX_URLS:
                capped = len(urls) - MAX_URLS
                urls = urls[:MAX_URLS]

            # ── Pre-flight info banners ─────────────────────────────
            # Build the list of side-banner messages dynamically
            warning_msgs = []
            if dupes:
                warning_msgs.append(f"**{dupes}** duplicate{'s' if dupes != 1 else ''} removed")
            if skipped_tld:
                # Group by TLD for a clean summary
                by_tld: dict[str, int] = {}
                for _, t in skipped_tld:
                    by_tld[t] = by_tld.get(t, 0) + 1
                tld_summary = ", ".join(f"{n} `{t}`" for t, n in sorted(by_tld.items()))
                warning_msgs.append(f"**{len(skipped_tld)}** non-commercial skipped ({tld_summary})")
            if skipped_chinese_urls:
                warning_msgs.append(f"**{len(skipped_chinese_urls)}** Chinese `.cn` skipped")
            if capped:
                warning_msgs.append(f"**{capped}** truncated (max {MAX_URLS})")

            info_cols = st.columns(3 if warning_msgs else 2)
            with info_cols[0]:
                st.info(f"📋 **{len(urls)}** unique URL{'s' if len(urls) != 1 else ''} to scrape")
            secs, domains = _estimate_time(urls)
            with info_cols[1]:
                mins = secs // 60
                time_str = f"~{mins}m {secs % 60}s" if mins else f"~{secs}s"
                st.info(f"⏱️ Estimated time: **{time_str}** across {domains} domain{'s' if domains != 1 else ''}")
            if warning_msgs:
                with info_cols[2]:
                    st.warning("ℹ️ " + " · ".join(warning_msgs))

            # If filtering left us with nothing, stop here
            if not urls:
                st.error("All URLs were filtered out. Untick *Skip non-commercial domains* in the sidebar to include them.")
                st.stop()

            # Create the DB row first — every URL completion writes back to it
            sid = db.create_scrape(_get_db(), list(urls), source_name, bool(skip_chinese))
            st.session_state.scrape_state = {
                "id":           sid,
                "total":        len(urls),
                "done_count":   0,
                "remaining":    list(urls),
                "results":      [],
                "failed":       [],
                "source_name":  source_name,
                "skip_chinese": bool(skip_chinese),
            }

            _run_scrape(list(urls))

            if not st.session_state.scrape_state["remaining"]:
                _finalize_scrape()
                st.rerun()


# ── Results ────────────────────────────────────────────────────────
results = st.session_state.results
failed  = st.session_state.failed

if results:
    st.markdown("---")
    # Stats
    total_emails = sum(len([e for e in (r["emails"] or "").split(",") if e.strip()]) for r in results)
    total_phones = sum(len([p for p in (r["phones"] or "").split(",") if p.strip()]) for r in results)
    total_cities = sum(1 for r in results if r["city"])

    total_companies = sum(1 for r in results if r.get("company"))
    total_forms     = sum(1 for r in results if r.get("contact_form"))
    total_people    = sum(1 for r in results if r.get("people"))
    total_niches    = len({r.get("niche") for r in results if r.get("niche")})

    m1, m2, m3, m4, m5, m6, m7, m8 = st.columns(8)
    m1.metric("Sites",      len(results))
    m2.metric("Emails",     total_emails)
    m3.metric("Phones",     total_phones)
    m4.metric("Companies",  total_companies)
    m5.metric("Niches",     total_niches)
    m6.metric("Key people", total_people)
    m7.metric("Cities",     total_cities)
    m8.metric("Has form",   total_forms)

    # Filter row — text search + niche dropdown
    fc1, fc2 = st.columns([3, 1])
    with fc1:
        q = st.text_input("🔎 Filter by URL, email, phone, social, or city",
                          value="", key="filter")
    with fc2:
        all_niches = sorted({r.get("niche") for r in results if r.get("niche")})
        niche_pick = st.selectbox(
            "Niche",
            options=["All"] + all_niches,
            key="niche_filter",
        )

    # Normalize results — make sure every row has every column, even if the result
    # was scraped before a new column (like "tech" or "niche") was added
    ALL_FIELDS = ["url", "company", "niche", "people", "emails", "phones",
                  "socials", "city", "language", "tech", "contact_form"]
    norm_results = [{f: r.get(f, "") for f in ALL_FIELDS} for r in results]
    df = pd.DataFrame(norm_results)

    if q:
        ql = q.lower()
        df = df[df.apply(lambda row: any(ql in str(v).lower() for v in row.values), axis=1)]
    if niche_pick != "All":
        df = df[df["niche"] == niche_pick]

    # Friendly column names & order
    df = df.rename(columns={
        "url":          "Website",
        "company":      "Company",
        "niche":        "Niche",
        "people":       "Key People",
        "emails":       "Emails",
        "phones":       "Phones",
        "socials":      "Socials",
        "city":         "City",
        "language":     "Lang",
        "tech":         "Built With",
        "contact_form": "Has form",
    })
    preferred = ["Website", "Company", "Niche", "Key People", "Emails", "Phones",
                 "City", "Built With", "Lang", "Has form", "Socials"]
    df = df[[c for c in preferred if c in df.columns]]

    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Website": st.column_config.LinkColumn(
                "Website",
                display_text=r"^(?:https?://)?(.+)$",
            ),
        },
    )

    # Downloads — name file after source (Google Sheet title or uploaded filename)
    base_name = _safe_filename(st.session_state.source_name) if st.session_state.source_name else "scrape-results"

    csv_buf = io.StringIO()
    df.to_csv(csv_buf, index=False)
    xlsx_buf = io.BytesIO()
    with pd.ExcelWriter(xlsx_buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Results", index=False)

    if st.session_state.source_name:
        st.caption(f"📁 Downloads will be named after: **{st.session_state.source_name}**")

    d1, d2, d3 = st.columns([1, 1, 4])
    with d1:
        st.download_button(
            "⬇ CSV", data=csv_buf.getvalue(),
            file_name=f"{base_name}.csv", mime="text/csv",
            use_container_width=True,
        )
    with d2:
        st.download_button(
            "⬇ Excel", data=xlsx_buf.getvalue(),
            file_name=f"{base_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    # Copy-all emails (display as a code block for easy copy)
    with st.expander("📋 Copy all emails as one string"):
        all_emails = ", ".join(r["emails"] for r in results if r["emails"])
        st.code(all_emails or "(none)")


# ── Failed list ────────────────────────────────────────────────────
if failed:
    with st.expander(f"⚠ Skipped — {len(failed)} URL(s) had no email"):
        for u in failed:
            st.markdown(f"- ❌ `{u}`")


# ── Sidebar: history ───────────────────────────────────────────────
with st.sidebar:
    # ── Public usage counter ─────────────────────────────────────
    total = _get_counter()
    if total is not None:
        st.markdown(
            f"<div style='padding:.6rem .85rem; background:#eff6ff; border:1px solid #bfdbfe; "
            f"border-radius:.5rem; margin-bottom:1rem; text-align:center;'>"
            f"<div style='font-size:1.4rem; font-weight:800; color:#1d4ed8;'>{total:,}</div>"
            f"<div style='font-size:.7rem; color:#64748b; text-transform:uppercase; letter-spacing:.05em;'>"
            f"total scrapes worldwide</div></div>",
            unsafe_allow_html=True,
        )

    st.markdown("### 🕘 History")

    # ── Download / Restore history ──────────────────────────────
    with st.expander("💾 Save / Restore"):
        # Save
        if st.session_state.history:
            hist_json = json.dumps(
                st.session_state.history,
                indent=2, default=str, ensure_ascii=False,
            )
            st.download_button(
                "⬇ Download history (.json)",
                data=hist_json,
                file_name=f"scraper-history-{datetime.now().strftime('%Y%m%d-%H%M')}.json",
                mime="application/json",
                use_container_width=True,
            )
        else:
            st.caption("Nothing to save yet.")

        # Restore
        restored_file = st.file_uploader(
            "Restore from .json",
            type=["json"],
            key="hist_restore",
            label_visibility="collapsed",
        )
        if restored_file is not None:
            try:
                loaded = json.loads(restored_file.read())
                if isinstance(loaded, list) and all(isinstance(h, dict) for h in loaded):
                    # Merge: new history first, dedup by 'when' timestamp
                    existing_keys = {h.get("when") for h in st.session_state.history}
                    incoming = [h for h in loaded if h.get("when") not in existing_keys]
                    st.session_state.history = (incoming + st.session_state.history)[:HIST_MAX * 4]
                    st.success(f"Restored {len(incoming)} new entries (total: {len(st.session_state.history)})")
                else:
                    st.error("Invalid history file format.")
            except json.JSONDecodeError:
                st.error("Could not parse JSON file.")
            except Exception as e:
                st.error(f"Could not restore: {e}")

    # History is now read from the SQLite DB so it survives container restarts
    # and is shared across browser tabs / sessions on the same instance.
    db_history = []
    try:
        db_history = db.get_history(_get_db(), limit=20)
    except Exception as exc:
        st.caption(f"⚠ DB read failed: {exc}")

    if not db_history:
        st.caption("No previous scrapes yet.")
    else:
        for h in db_history:
            with st.container(border=True):
                source = h.get("source_name") or "(direct input)"
                when_short = (h.get("started_at") or "")[:16].replace("T", " ")
                status_html = ""
                if not h.get("finished_at"):
                    status_html = (
                        " · <span style='color:#b45309'><b>unfinished</b></span>"
                    )
                st.markdown(
                    f"<small style='color:#1e293b'><b>{source}</b></small><br>"
                    f"<span style='font-size:.85rem;'>"
                    f"**{h['done']}** done · "
                    f"{('**' + str(h['failed']) + '** failed · ') if h['failed'] else ''}"
                    f"{('**' + str(h['pending']) + '** pending' + status_html) if h['pending'] else ''}"
                    f"</span><br>"
                    f"<small style='color:#94a3b8'>{when_short}</small>",
                    unsafe_allow_html=True,
                )
                c1, c2 = st.columns(2)
                if c1.button("Open", key=f"hist_open_{h['id']}", use_container_width=True):
                    conn = _get_db()
                    st.session_state.results     = db.get_results(conn, h["id"])
                    st.session_state.failed      = db.get_failed_urls(conn, h["id"])
                    st.session_state.source_name = h.get("source_name")
                    st.rerun()
                if c2.button("Delete", key=f"hist_del_{h['id']}", use_container_width=True):
                    db.delete_scrape(_get_db(), h["id"])
                    # If the deleted row was the active scrape, clear the banner too
                    if (st.session_state.scrape_state
                        and st.session_state.scrape_state.get("id") == h["id"]):
                        st.session_state.scrape_state = None
                    st.rerun()

    st.markdown("---")
    st.markdown(
        """
        <div style="font-size:.78rem; color:#64748b; line-height:1.5;">
          <div style="margin-bottom:.4rem;">
            🤖 <b>Respects robots.txt</b> — disallowed URLs silently skipped.
          </div>
          <div style="margin-bottom:.4rem;">
            🐢 <b>Per-domain rate limited</b> — 1 req/sec per domain.
          </div>
          <div style="margin-bottom:.4rem;">
            🧹 <b>Tracking params stripped</b> — <code>utm_*</code>, <code>fbclid</code>, etc.
          </div>
          <div style="margin-bottom:.4rem;">
            🛡 <b>Auto-retries</b> with browser headers on 403 / Cloudflare.
          </div>
          <div>
            📄 Only scrapes <b>homepage</b>, <b>contact</b>, <b>about</b> pages.
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
